from SOAPpy import WSDL
from ConfigParser import SafeConfigParser
from openpyxl import load_workbook
import subprocess
import json
from xml.etree.ElementTree import ElementTree, Element, SubElement
from threading import Thread, Lock


class Crawler():
    def __init__(self):
        self.parseConfigFile()

    wsdlUrl = ''
    urlCategories = {}

    email = ''
    password = ''
    urlCategoryXmlPath = ''
    feedPath = ''
    phpPath = ''
    amountOfThreads = ''

    client = ''
    sessionHash = ''

    amountOfThreads = ''
    activeThreads = 0

    lock = Lock()
    
    def main(self):
        self.getUrlCategories()
        
        self.client = WSDL.Proxy(self.wsdlUrl)
        self.sessionHash = self.client.login(self.email, self.password).Records.item.value
        
        self.gatherData()

    # Procedure to parse the config file
    def parseConfigFile(self):
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')
        self.email = parser.get('General', 'email')
        self.password = parser.get('Belboon', 'password')
        self.urlCategoryXmlPath = parser.get('Belboon', 'urlcategoryxml')
        self.phpPath = parser.get('Belboon', 'phppath')
        self.feedPath = parser.get('Belboon', 'feedpath')
        self.wsdlUrl = parser.get('Belboon', 'wsdlurl')
        self.amountOfThreads = int(parser.get('General', 'amountofthreads'))
          
    # Some campaigns return multiple feeds. The url category xml contains which feed to get, so load this
    def getUrlCategories(self):
        wb = load_workbook(filename=self.urlCategoryXmlPath, use_iterators=True)
        ws = wb.get_sheet_by_name('Sheet1')
        
        for row in ws.iter_rows(row_offset=1):
            for cell in row:
                if cell.column == 'A':  # Website url
                    website = cell.value
                elif cell.column == 'B':  # Correct feed category
                    self.urlCategories[website] = cell.value
    
    # Gather feed IDs and call php script to download product data for these feed id's.
    def gatherData(self):
        feeds = self.client.getFeeds(self.sessionHash)
        
        feedUrl = ''
        passedElektro = False
        passedKabel = False
        for feed in feeds.Records:
            # Somehow, these are passed twice.
            if feedUrl == 'www.elektro2000.de' and passedElektro is False:
                passedElektro = True
                continue
            if feedUrl == 'www.kabelmeister.de' and passedKabel is False:
                passedKabel = True
                continue

            feedId = feed.item[0].value
            feedUrl = self.stripUrl(feed.item[5].value)
            
            # Check if the current campaign contains multiple feeds and if so, check what feed to pick.
            try:
                xmlCategory = self.urlCategories[feedUrl]
                feedCategory = feed.item[1].value
                
                if xmlCategory != feedCategory:
                    continue
            except KeyError:
                pass

            #  Use threads for parsing product data to XML files.
            while True:
                if self.activeThreads < self.amountOfThreads:
                    self.lock.acquire()
                    try:
                        t = Thread(target=self.writeXML, args=(feedUrl, feedId))
                        t.start()

                        print 'Crawling ' + feedUrl

                        self.activeThreads += 1
                    finally:
                        self.lock.release()
                    break

    # Cleans and transforms the url to the required format
    def stripUrl(self, feedUrl):
        # Remove 'http://' and the slash if it's the last character.
        if feedUrl.find('http://') != -1:
            feedUrl = feedUrl[7:]
        if feedUrl[-1:] == '/':
            feedUrl = feedUrl[0:-1]
            
        # Check if the url contains anything other then a country extension after the '/'
        lastDot = feedUrl.rfind('.')
        lastSlash = feedUrl[lastDot:].find('/')
        if lastSlash != -1 and len(feedUrl[lastSlash + lastDot + 1:]) > 3:
            feedUrl = feedUrl[:lastSlash + lastDot]
            
        return feedUrl.replace('/', '$')  # Can't save a file with a slash, this will be re-done by BOB.'

    # Procedure to convert to XML and save it as a xml file.
    def writeXML(self, feedUrl, feedId):
        # Call the php script that takes feedId as an argument and downloads the product data, which
        # is returned in JSON format.
        try:
            result = json.loads(subprocess.check_output(["php", self.phpPath, str(feedId)]).decode('utf-8'))

            root = Element('products')

            for record in result['Records']:
                product = SubElement(root, 'product')
                for key in record:
                    child = SubElement(product, key)
                    child.text = record[key]

            ElementTree(root).write(self.feedPath + feedUrl + '.xml', encoding='UTF-8')
        except Exception as e:
            print e

        self.lock.acquire()
        try:
            self.activeThreads -= 1
        finally:
            self.lock.release()

        print 'Done crawling ' + feedUrl