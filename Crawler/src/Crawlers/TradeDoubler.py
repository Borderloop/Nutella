import urllib2
import urllib
import time
import traceback
import re
from openpyxl import load_workbook
from ConfigParser import SafeConfigParser

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET


from CrawlerHelpScripts import logger


class Crawler():
    def __init__(self):
        self.parseConfigFile()

    token = ''
    nameUrlsPath = ''
    feedPath = ''
    
    log = logger.createLogger("TradeDoublerLogger", "TradeDoubler")
    
    prevName = ''
    x = 2
    name = ''

    # Keep track if the name is found. Default is true so names that are by default correct urls will get saved.
    found = True
    
    def main(self):
        start_time = time.time()
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Starting with TradeDoubler")
        
        #Get token 
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')
        self.token = parser.get('TradeDoubler', 'token')
        
        #Add token to the url. This url links to a xml file that contains program data
        url = "http://api.tradedoubler.com/1.0/productFeeds.xml?token=%s" % self.token
        xmlfile = urllib2.urlopen(url)

        #Create a string from the data
        data = xmlfile.read()
        xmlfile.close()
        root = ET.fromstring(data)

        #Each child is an affiliate, so for each one get feedID and na
        for record in root:
            for child in record:
                self.gatherData(child)
        
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Finished crawling TradeDoubler in: " + str((time.time() - start_time)))

    # This procedure parses the config file
    def parseConfigFile(self):
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')
        self.token = parser.get('TradeDoubler', 'token')
        self.feedPath = parser.get('TradeDoubler', 'feedpath')
        self.nameUrlsPath = parser.get('General', 'nameurlspath')

    #Gathers the data needed for downloading and correct file name            
    def gatherData(self, child):
        #Extract feedId, name and numberOfProducts from xml file
        if child.tag == "{urn:com:tradedoubler:pf:model:xml:common}feedId":
            self.feedId = child.text
            
        if child.tag == "{urn:com:tradedoubler:pf:model:xml:common}name":
            self.name = child.text
            #Check if the name contains a domain extension. If not, it means the name is not an url. Change this
            regexp = re.compile(r'\.[a-z]{2,3}')
            if regexp.search(self.name) is None:
                self.changeNameInURL()
            
            #If the previous name equals the current name, this is a nth file of the same company.
            # Add a number to the end to make the filename unique and so it doesnt get overwritten.
            if self.prevName == self.name:
                self.log.info("nth child for: " + self.name)
                self.name = self.name + " " + str(self.x)
                self.x += 1
            else:
                self.x = 2  # Always reset x if a new company is found

                # If the name is a correct url, start downloading and saving file.
                if self.found is True:
                    self.save()

    #Downloads and saves the xml file under the correct name      
    def save(self):
        print 'Saving ' + self.name
        feedURL = "http://api.tradedoubler.com/1.0/productsUnlimited.xml;fid=%s?token=%s" % (self.feedId, self.token)

        #If the save fails, something is wrong with the file or directory name. Catch this error
        try:
            xmlFile = urllib.URLopener()
            xmlFile.retrieve(feedURL, "C:\BorderSoftware\Boris\ProductFeeds\TradeDoubler" + self.name + ".xml")
            print 'Done saving ' + self.name
        except Exception as e:
            self.log.error(str(time.asctime(time.localtime(time.time()))) + ": " + traceback.format_exc())
            self.log.info(str(time.asctime(time.localtime(time.time()))) + ": Failed:" + str(e))
        
        self.prevName = self.name    
        
    #This method changes the name into an url, by looking in an XML file for the correct url for a given name
    def changeNameInURL(self):
        self.found = False
        
        #Import the excel file with names and associated urls
        wb = load_workbook(filename=self.nameUrlsPath, use_iterators=True)
        ws = wb.get_sheet_by_name('Replace')
        
        #Iterate trough all rows
        for row in ws.iter_rows(row_offset=1):
            for cell in row:
                if cell.column == 'A':  # Column A contains the name as given by TradeDoubler

                    # If the cell contains the given name, change it to the value stored in the column after (B),
                    # which is the correct url.
                    if cell.value == self.name: 
                        self.name = ws['B' + str(cell.row)].value
                        self.found = True
        
        # If the name is not found, the name is not yet inserted in the nameurls.xlsx file.
        # This needs to be added. Log this so it's easily traced.
        if self.found is False:
            self.log.error(str(time.asctime(time.localtime(time.time()))) + ": NAME NOT FOUND!!! Didn't find the name '"+ self.name +"' in the nameurls file. Add this")