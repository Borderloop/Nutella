from openpyxl import load_workbook, Workbook
import urllib
from CrawlerHelpScripts import logger
import time

class Crawler():
    
    log = logger.createLogger("ZanoxLogger", "Zanox")
    
    #Import the excel file with websites
    wb=load_workbook(r'C:\feed_urls.xlsx', use_iterators = True)
    ws=wb.get_sheet_by_name('Sheet1')
    
    def main(self):
        start_time = time.time()
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Starting with Zanox")
        
        self.gatherData()
        
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Finished crawling Zanox in: " + str((time.time() - start_time)))
    
    #Gathers the data needed for downloading and correct file name    
    def gatherData(self):
        #Iterate trough all rows
        for row in self.ws.iter_rows(row_offset=1):
            for cell in row:
                if cell.column == 'B': #Column B contains the name of the website.
                    website = cell.value
                elif cell.column == 'C': #Else if column is C, it contains the product feed download url. 
                    url = cell.value
                    self.save(url, website)
            
    #Downloads and saves the xml file under the correct name           
    def save(self, url, website):
        #Download xml file and write it to a file
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Saving xml file for: " + website )
        
        xmlFile = urllib.URLopener()
        xmlFile.retrieve(url, "C:/Product Feeds/Zanox/" + website + ".xml")
        
        self.log.info(str(time.asctime( time.localtime(time.time()) ))+": Done saving xml file for: " + website )
        