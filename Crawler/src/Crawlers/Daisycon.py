from CrawlerHelpScripts import logger
import suds
from suds.transport.http import HttpAuthenticated
from suds.xsd.doctor import ImportDoctor, Import
import hashlib
from ConfigParser import SafeConfigParser
import subprocess
import json
import urllib
import re
from openpyxl import load_workbook
import time
import traceback
from threading import Thread, Lock


class Crawler():
    def __init__(self):
        self.parseConfigFile()

    wsdlUrl = ''
    email = ''
    password = ''
    targetNameSpace = ''
    phpPath = ''
    feedPath = ''
    nameUrlsPath = ''

    amountOfThreads = ''
    activeThreads = 0

    client = ''

    regex = '^[a-zA-Z0-9\-\.]+\.(com|org|net|mil|nl|fr|pl|it|de|be|uk|co\.uk|us)/?(com|org|net|mil|nl|fr|pl|it|de|be|uk|co\.uk|us)?'  # url validation regular expression.

    log = logger.createLogger("DaisyconLogger", "Daisycon")

    lock = Lock()

    def main(self):
        #Login first
        md5pass = hashlib.md5(self.password).hexdigest()
        credentials = dict(username=self.email, password=md5pass)
        t = HttpAuthenticated(**credentials)

        imp = Import('http://schemas.xmlsoap.org/soap/encoding/')
        imp.filter.add('http://api.daisycon.com/publisher/soap//program/')
        d = ImportDoctor(imp)
        self.client = suds.client.Client(self.wsdlUrl, doctor=d, transport=t)
        self.gatherData()

    # Procedure to parse the config file
    def parseConfigFile(self):
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')
        self.email = parser.get('General', 'email')
        self.password = parser.get('General', 'password')
        self.wsdlUrl = parser.get('Daisycon', 'programwsdlurl')
        self.targetNameSpace = parser.get('Daisycon', 'targetnamespace')
        self.phpPath = parser.get('Daisycon', 'phppath')
        self.feedPath = parser.get('Daisycon', 'feedpath')
        self.nameUrlsPath = parser.get('General', 'nameurlspath')
        self.amountOfThreads = int(parser.get('General', 'amountofthreads'))

    # Gets the program id's for approved programs, then calls php script that gets the feed download url.
    def gatherData(self):

        filter = dict(status='active')
        subscriptions = self.client.service.getSubscriptions(filter)

        for subscription in subscriptions[0]:
            try:
                if subscription.media[0].status == 'approved':  # Disapproved programs still exist here despite filter.
                    programId = subscription.program_id

                    # Call the php script that retrieves the name and download link for the xml file.
                    try:
                        result = json.loads(subprocess.check_output(["php", self.phpPath, str(programId)]).decode('utf-8'))
                    except ValueError:
                        continue

                    self.validateWebsiteUrl(result)
            except Exception as e:
                print e

    # Procedure to check if a given url is correct and if not, try to make it so.
    def validateWebsiteUrl(self, result):

        websiteUrl = 'www.' + result['websiteName'].lower()

        # Check if it's a valid url by regular expression
        if re.search(self.regex, websiteUrl):
            self.startThread(websiteUrl, result['feedUrl'])
        else:
            save = False

            wb = load_workbook(filename=self.nameUrlsPath, use_iterators=True)
            ws = wb.get_sheet_by_name('Replace')

            for row in ws.iter_rows(row_offset=1):
                for cell in row:
                    # If the website name is found here, it needs a replace,
                    # which is located in the B column for that row.
                    if cell.column == 'A' and result['websiteName'].lower() in cell.value:
                        websiteUrl = ws['B' + str(cell.row)].value
                        self.startThread(websiteUrl, result['feedUrl'])
                        save = True

            if save is False:
                ws = wb.get_sheet_by_name('Daisycon')

                # If it still didn't pass yet, it might be because of a missing extension. Check for this.
                for row in ws.iter_rows(row_offset=1):
                    for cell in row:
                        if cell.column == 'A' and result['websiteName'].lower() in cell.value:
                            websiteUrl = websiteUrl + ws['B' + str(cell.row)].value
                            self.startThread(websiteUrl, result['feedUrl'])
                            save = True

            if save is False:
                self.log.error(str(time.asctime(time.localtime(time.time()))) + ": Not a valid url: " + websiteUrl)
                return False

    # Starts a thread to save feed.
    def startThread(self, websiteURL, feedURL):
        while True:
            if self.activeThreads < self.amountOfThreads:
                self.lock.acquire()
                try:
                    t = Thread(target=self.save, args=(websiteURL, feedURL))
                    t.start()

                    self.activeThreads += 1
                finally:
                    self.lock.release()
                break

    # Downloads and saves the xml file under the correct name
    def save(self, websiteURL, feedURL):
        print 'Crawling ' +websiteURL
        websiteUrl = websiteURL.replace('/', '$')
        # If the save fails, something is wrong with the file or directory name. Catch this error
        try:
            xmlFile = urllib.URLopener()
            xmlFile.retrieve(feedURL, self.feedPath + websiteUrl + ".xml")
            print 'Done crawling ' + websiteUrl
        except:
            self.log.error(str(time.asctime(time.localtime(time.time()))) + ": " + traceback.format_exc())
            self.log.info(str(time.asctime(time.localtime(time.time()))) + ": Failed:" + websiteUrl)

        self.lock.acquire()
        try:
            self.activeThreads -= 1
        finally:
            self.lock.release()