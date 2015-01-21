# This module crawls multiple affiliates, stored in the feed_urls.xlsx file

from CrawlerHelpScripts import logger

from openpyxl import load_workbook
import urllib
import time
import traceback
from ConfigParser import SafeConfigParser
import zipfile
import os
from threading import Thread, Lock


class Crawler():
    def __init__(self):
        self.parseConfigFile()

    feedUrlsPath = ''
    feedPath = ''

    log = logger.createLogger("FeedCrawlerLogger", "FeedCrawler")
    
    #Import the excel file with websites
    wb = ''
    ws = ''

    amountOfThreads = ''
    activeThreads = 0
    
    prevName = ""  # Used to check if the previous campaign is from the same company.
    x = 2  # Used to name files of companies that use multiple xml files

    lock = Lock()
    
    def main(self):
        start_time = time.time()
        self.log.info(str(time.asctime(time.localtime(time.time())))+": Starting with FeedCrawler")

        #Import the excel file with websites
        self.wb = load_workbook(filename=self.feedUrlsPath, use_iterators=True)
        self.ws = self.wb.get_sheet_by_name('Sheet1')

        self.gatherData()
        
        self.log.info(str(time.asctime(time.localtime(time.time()))) +
                      ": Finished crawling FeedCrawler affiliates in: " + str((time.time() - start_time)))

    # Procedure to parse the config file
    def parseConfigFile(self):
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')
        self.feedUrlsPath = parser.get('FeedCrawler', 'feedurlspath')
        self.feedPath = parser.get('FeedCrawler', 'feedpath')
        self.amountOfThreads = int(parser.get('General', 'amountofthreads'))

    #Gathers the data needed for downloading and correct file name    
    def gatherData(self):
        #Iterate trough all rows
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.column == 'A':  # Column A contains the affiliate name.
                    affiliate = cell.value
                if cell.column == 'B':  # Column B contains the name of the website.
                    website = cell.value
                if cell.column == 'C':  # Column C contains the file type (xml or zip)
                    fileType = cell.value
                if cell.column == 'D':  # Else if column is D, it contains the product feed download url.
                    url = cell.value

                    # If there's a semicolon, there are multiple product feeds for this company,
                    # so save them all and add a number.
                    if url.find(';') != -1:
                        urls = url.split(';')
                        x = 1
                        for url in urls:
                            websiteWithNumber = website + " " + str(x)
                            print 'Crawling ' +websiteWithNumber
                            self.startThread(affiliate, url, websiteWithNumber, fileType)
                            x = x+1
                    else:  # Else it's just one url, save this
                        print 'Crawling ' +website
                        self.startThread(affiliate, url, website, fileType)

    # Starts a thread to save feed.
    def startThread(self, affiliate, url, website, fileType):
        while True:
            if self.activeThreads < self.amountOfThreads:
                self.lock.acquire()
                try:
                    t = Thread(target=self.save, args=(affiliate, url, website, fileType))
                    t.start()

                    self.activeThreads += 1
                finally:
                    self.lock.release()
                break

    # Downloads and saves the xml file under the correct name
    def save(self, affiliate, url, website, fileType):
        website = website.replace('/', '$')
        # Download xml file and write it to a file
        # If the save fails, something is wrong with the file or directory name. Catch this error
        try:
            xmlFile = urllib.URLopener()
            xmlFile.retrieve(url, self.feedPath + affiliate + "/" + website + '.' + fileType)
            xmlFile.close()

            if fileType == 'zip':
                self.readZipFile(website, affiliate)

            print 'Done crawling ' + website

        except:
            self.log.error(str(time.asctime(time.localtime(time.time()))) +
                           ": " + traceback.format_exc())
            self.log.info(str(time.asctime(time.localtime(time.time())))+": Failed saving file for: " + affiliate +
                          " - " + website + ". See error log for more details")

        self.lock.acquire()
        try:
            self.activeThreads -= 1
        finally:
            self.lock.release()

    # This procedure is used to read zip files and extract their contents and afterwards delete the zip file.
    def readZipFile(self, website, affiliate):
        zf = zipfile.ZipFile(self.feedPath + affiliate + '/' + website + '.zip')
        nameList = zf.namelist()
        data = zf.read(nameList[0])
        zf.close()

        f = open(self.feedPath + affiliate + '/' + website + '.csv', 'w')
        f.write(data)
        f.close()

        os.remove(self.feedPath + affiliate + '/' + website + '.zip')