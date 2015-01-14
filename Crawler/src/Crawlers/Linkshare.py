import urllib
import time
import gzip
import os
import traceback

from ConfigParser import SafeConfigParser
from openpyxl import load_workbook
from threading import Thread, Lock

from CrawlerHelpScripts import logger


class Crawler():
    def __init__(self):
        self.parseConfigFile()

    feedPath = ''
    xmlPath = ''
    user = ''
    password = ''
    sid = ''

    amountOfThreads = ''
    activeThreads = 0

    log = logger.createLogger("LinkshareLogger", "Linkshare")

    lock = Lock()

    # This procedure controls the main flow of the program.
    def main(self):
        self.gatherData()

    # Procedure to parse the config file
    def parseConfigFile(self):
        parser = SafeConfigParser()
        parser.read('C:/BorderSoftware/Boris/settings/boris.ini')

        self.feedPath = parser.get('Linkshare', 'feedpath')
        self.xmlPath = parser.get('Linkshare', 'xmlpath')
        self.user = parser.get('Linkshare', 'user')
        self.password = parser.get('Linkshare', 'password')
        self.sid = parser.get('Linkshare', 'sid')
        self.amountOfThreads = int(parser.get('General', 'amountofthreads'))

    # This procedure gathers the advertiser id and website url, needed for downloading and saving
    # the file.
    def gatherData(self):

        wb = load_workbook(filename=self.xmlPath, use_iterators=True)
        ws = wb.get_sheet_by_name('Sheet1')

        for row in ws.iter_rows(row_offset=1):
            for cell in row:
                # Affiliate id is in column A, the corresponding url in B. Get both and save it
                if cell.column == 'A':

                    try:
                        affiliateID = str(int(cell.value))
                        websiteURL = ws['B' + str(cell.row)].value
                    except TypeError:
                        break

                    if 'rakuten' not in websiteURL:
                        while True:
                            if self.activeThreads < self.amountOfThreads:
                                self.lock.acquire()
                                try:
                                    t = Thread(target=self.save, args=(affiliateID, websiteURL))
                                    t.start()

                                    self.activeThreads += 1
                                finally:
                                    self.lock.release()
                            break

    # Downloads and saves the files under the correct name
    def save(self, affiliateID, websiteURL):
        print 'Crawling ' + websiteURL

        # If the save fails, something is wrong with the file or directory name. Catch this error
        tries = 0
        while True:
            try:
                zipFile = urllib.URLopener()

                zipFile.retrieve('ftp://' + self.user + ':' + self.password + '@aftp.linksynergy.com/' + affiliateID + '_' +
                                 self.sid + '_mp.txt.gz', self.feedPath + websiteURL + ".txt.gz")

                self.readZipFile(websiteURL)
                print 'Done crawling ' + websiteURL

                break
            except IOError:
                tries += 1
                time.sleep(1)
                if tries == 5:
                    break
            except Exception as e:
                self.log.error(str(time.asctime(time.localtime(time.time()))) + ": " + str(e))
                self.log.info(str(time.asctime(time.localtime(time.time()))) + ": Failed:" + websiteURL)

                break

        self.lock.acquire()
        try:
            self.activeThreads -= 1
        finally:
            self.lock.release()

    # This procedure is used to read zip files and extract their contents and afterwards delete the zip file.
    def readZipFile(self, website):
        zipFile = gzip.open(self.feedPath + website + '.txt.gz')
        csvFile = open(self.feedPath + website + '.csv', 'a')

        for line in zipFile:
            csvFile.write(line)

        zipFile.close()
        csvFile.close()

        os.remove(self.feedPath + website + '.txt.gz')